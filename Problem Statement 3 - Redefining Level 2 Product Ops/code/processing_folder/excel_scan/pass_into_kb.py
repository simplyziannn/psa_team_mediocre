import io, sys, re, json
from typing import Any, Dict, Optional, Tuple, Union
from contextlib import redirect_stdout
from pathlib import Path
from openpyxl import load_workbook

# ---- Ensure we can import from project root (…/code) ----
PROJECT_ROOT = Path(__file__).resolve().parents[2]  # .../code
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

# ---- Project imports ----
from email_processor.email_processor import process_any
from processing_folder.excel_scan.excel_scanner import check_excel_for_string
from processing_folder.excel_scan.extract_text_sample import extract_text_sample  # returns (query, source)

Field = Union[str, int]

# --------------------------------------------------------------------
# Helper: parse top match from the scanner's printed output
# --------------------------------------------------------------------
def get_highest_confidence_from_printed(printed_output: str, debug: bool = False) -> Optional[Dict[str, Any]]:
    """
    Extracts the highest-confidence match from the printed scanner output.

    Looks for the first occurrence like: "[SheetName r323 c7]" after "✅ Found matches:".
    Returns: {"sheet": str, "row": int, "col": int, "score": float} or None.
    """
    marker = "✅ Found matches:"
    if marker not in printed_output:
        if debug:
            print("❌ No 'Found matches' section in output.")
        return None

    after_marker = printed_output.split(marker, 1)[-1].strip()

    _SCORE_RE = re.compile(r"(?P<score>-?\d+(?:\.\d+)?)")
    _LOCATOR_RE = re.compile(r"\[(?P<sheet>[^\]]+?)\s+r(?P<row>\d+)\s+c(?P<col>\d+)\]")

    mloc = _LOCATOR_RE.search(after_marker)
    mscore = _SCORE_RE.search(after_marker)

    if not mloc:
        if debug:
            print("❌ Could not locate any [Sheet r c] pattern.")
        return None

    sheet = mloc.group("sheet").strip()
    row = int(mloc.group("row"))
    col = int(mloc.group("col"))
    score = float(mscore.group("score")) if mscore else None

    result = {"sheet": sheet, "row": row, "col": col, "score": score}
    if debug:
        print(f"🏆 Top match -> Sheet: {sheet}, Row: {row}, Col: {col}, Score: {score}")
    return result


# --------------------------------------------------------------------
# Helper: read F/G/H from a given sheet+row
# --------------------------------------------------------------------
def read_fixed_columns_from_excel(xlsx_path: str, sheet: str, row: int, debug: bool = False) -> Tuple[Field, Field, Field, bool]:
    """
    Reads columns F(6), G(7), H(8) for a given (sheet, row).
    Assumes the 'row' is 0-based from the scanner; converts to 1-based for Excel.
    Returns: (col6, col7, col8, match_status)
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        ws = wb.active
        if debug:
            print(f"⚠️ Sheet '{sheet}' not found, using first sheet instead.")
    else:
        ws = wb[sheet]

    excel_row = row + 1 if row >= 0 else row

    values = []
    for col in (6, 7, 8):  # F/G/H
        cell_value = ws.cell(row=excel_row, column=col).value
        values.append(cell_value if (cell_value is not None and str(cell_value).strip() != "") else 0)

    match_status = any(v != 0 for v in values)

    if debug:
        coords = [ws.cell(row=excel_row, column=col).coordinate for col in (6, 7, 8)]
        print(f"📘 Reading from: {xlsx_path}")
        print(f"📄 Sheet: {ws.title}, Row: {excel_row}, Columns: F/G/H")
        print(f"🧭 Cells: {coords}")
        print(f"📖 Values: {values}")
        print(f"✅ Match found: {match_status}")

    return (*values, match_status)


# --------------------------------------------------------------------
# Public API: raw_email -> (F, G, H, matched)
# --------------------------------------------------------------------
def process_email_to_tuple(
    raw_email: str,
    excel_path: Optional[str] = None,
    debug: bool = False,
) -> Tuple[Field, Field, Field, bool]:

    """
    Pipeline:
      raw_email -> process_any(...) -> extract_text_sample(...) -> check_excel_for_string(query)
      -> parse top match -> read F/G/H on that row -> return tuple

    Returns:
      (F_value, G_value, H_value, matched: bool)
      or (0, 0, 0, False) if no usable query or no match.
    """
    if not raw_email or not raw_email.strip():
        return (0, 0, 0, False)

    # 1) Email -> structured result
    result = process_any(raw_email)
    if hasattr(result, "to_dict"):
        result = result.to_dict()
    if isinstance(result, str):
        try:
            result = json.loads(result)
        except json.JSONDecodeError:
            # keep as string; extractor will handle raw strings, but likely not useful for Excel
            pass

    # 2) Best query to search
    query, _ = extract_text_sample(result)
    if not query:
        return (0, 0, 0, False)

    # 3) Run Excel scanner and capture printed output
    buf = io.StringIO()
    with redirect_stdout(buf):
        check_excel_for_string(query)
    printed_output = buf.getvalue()

    # 4) Parse top match
    top = get_highest_confidence_from_printed(printed_output, debug=debug)
    if not top:
        return (0, 0, 0, False)

    # 5) Excel path (default to repo file)
    if excel_path is None:
        excel_path = str(PROJECT_ROOT / "processing_folder" / "excel_scan" / "Case Log.xlsx")

    # 6) Read F/G/H on that row
    return read_fixed_columns_from_excel(
        excel_path, sheet=top["sheet"], row=top["row"], debug=debug
    )


def main(raw_email):
    # 📨 Step 1: Provide the raw email content


    # Step 2: Run the pipeline and get the result
    result_tuple = process_email_to_tuple(raw_email, debug=False)
    #print(tuple(type(x).__name__ for x in result_tuple))

    return result_tuple
    # 🎯 Step 3: Display the results
    #print("\n🎯 Final Output Tuple:")
    print(result_tuple)

if __name__ == "__main__":
    raw_email = (
    "Email ALR-861600 | CMAU00000020 - Duplicate Container information received. "
    "Hi Jen, Please assist in checking container CMAU00000020. "
    "Customer on PORTNET is seeing 2 identical containers information."
    )
    main(raw_email)
